from openpyxl import Workbook, load_workbook

wb = load_workbook('original_tag_sheets\DCM_tag_sheets\Tags_2022 RAQC SSBA - Digital_RAQC (Explore HQ).xlsx')
ws = wb.active


# print(wb.sheetnames)
# print(ws.min_row)
# print(ws.max_row)
# print(ws.min_column)
# print(ws.max_column)


start_row = 13
end_row = 19
start_column = 6
end_column = 17


for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column, values_only=True):
    print(row)
    break
#     cell_values_in_row = []
#     for cell in row:
#         cell_values_in_row.append(cell.value)
#         print(cell.value)


placement_id_list = []
placement_name_list = []
javascript_tag_list = []

for col in ws.iter_cols(min_col=start_column, max_col=end_column, min_row=start_row, max_row=end_row, values_only=True):
    # print(col)
    # break

    if col[0] == 'Placement ID' or 'Placement Name' or 'JavaScript Tag':
        
        if col[0] == 'Placement ID':
            for cell in col:
                placement_id_list.append(cell)

        elif col[0] == 'Placement Name':
            for cell in col:
                placement_name_list.append(cell)

        elif col[0] == 'JavaScript Tag':
            for cell in col:
                javascript_tag_list.append(cell)

    else:
        pass
            

placement_id_dict = {}

print(placement_id_list)
print(placement_name_list)
print(javascript_tag_list)